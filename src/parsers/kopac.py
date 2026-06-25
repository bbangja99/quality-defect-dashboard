# -*- coding: utf-8 -*-
"""Kopac 공정 파서 (파일 2종 자동 판별).

(1) 불량유형 파레토 — 시트 'Defect Repair Summary By Type'
    좌(A~E)=1ml, 우(G~K)=3ml 의 [불량코드|불량명|불량수|점유율|누적점유율].
    불량명에 '공정검사' 포함 시 손실(loss)로 분류. → defect_detail 산출.
    (이 파일에는 투입수량이 없어 production 은 만들지 않음)

(2) 전수검사 — 시트 'Sheet1'
    제품(1ml/3ml) 블록마다 검사수량(투입)/적합/부적합(불량)/공정검사수량(손실)/유형별.
    검증: 부적합 = 검사 - 적합, 유형별 합 = 부적합, 공정검사는 부적합과 별도(=손실).
    → production + defect_detail(유형별) 산출.

주차: 두 파일 모두 본문에 주차가 없어 파일명 날짜(YYMMDD)로 주차를 정한다.
"""
from __future__ import annotations
import re

import openpyxl

from .base import BaseParser, ParseResult
from .. import normalize, config

_PARETO_SHEET = "Defect Repair Summary By Type"


def _strip_process_tag(name: str) -> str:
    """'콘 깨짐 (본딩)' → '콘 깨짐' : 끝의 (공정) 괄호 태그 제거."""
    return re.sub(r"\s*\([^)]*\)\s*$", "", str(name)).strip()


class KopacParser(BaseParser):
    process = "Kopac"

    def parse(self, file, week_label: str | None = None) -> ParseResult:
        # 주차 결정: 명시값 > 파일명 날짜
        year = week = week_num = None
        if week_label:
            year, week, week_num = normalize.normalize_week(week_label)
        else:
            name = getattr(file, "name", str(file))
            d = normalize.yymmdd_from_name(name)
            if d:
                year, week, week_num = normalize.week_from_date(d)

        wb = openpyxl.load_workbook(file, data_only=True)
        sheets = [s.strip() for s in wb.sheetnames]

        # 파일명으로 주차를 못 구하면 시트 내 '최신 날짜'로 추론(파레토가 다주차에 걸칠 때 최신주차로 라벨).
        inferred = False
        if week is None:
            d = self._latest_date(wb)
            if d:
                year, week, week_num = normalize.week_from_date(d)
                inferred = True

        if _PARETO_SHEET in sheets:
            res = self._parse_pareto(wb, year, week, week_num)
        else:
            res = self._parse_inspection(wb, year, week, week_num)
        if inferred:
            res.warnings.append(f"Kopac: 파일명 주차 해석 실패 → 시트 내 최신 날짜 기준 {week} 로 추론.")
        wb.close()
        return res

    @staticmethod
    def _latest_date(wb):
        """워크북 전체에서 가장 늦은 날짜 셀을 찾아 반환(주차 추론용)."""
        import datetime as _dt
        latest = None
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for v in row:
                    if isinstance(v, _dt.datetime):
                        d = v.date()
                        if latest is None or d > latest:
                            latest = d
        return latest

    # ── (1) 파레토 ─────────────────────────────────────────────────
    def _parse_pareto(self, wb, year, week, week_num) -> ParseResult:
        ws = wb[[s for s in wb.sheetnames if s.strip() == _PARETO_SHEET][0]]
        warnings: list[str] = []
        detail_rows = []
        # (시작열, 모델) : 1ml=A~, 3ml=G~ . 불량명=+1, 불량수=+2
        for col0, model in ((1, "1ml"), (7, "3ml")):
            r = 3
            while r <= ws.max_row:
                code = ws.cell(row=r, column=col0).value
                name = ws.cell(row=r, column=col0 + 1).value
                qty = ws.cell(row=r, column=col0 + 2).value
                if (code is None and name is None) or name is None:
                    break
                q = normalize.clean_number(qty)
                is_loss = normalize.is_loss_type("Kopac", "", defect_name=name)
                dtype = normalize.standardize_defect_type("Kopac", _strip_process_tag(name))
                detail_rows.append({
                    "year": year, "week": week, "week_num": week_num, "date": None,
                    "process": "Kopac", "model": model,
                    "defect_type": dtype, "defect_qty": q, "is_loss": is_loss,
                })
                r += 1
        if week is None:
            warnings.append("Kopac 파레토: 파일명에서 주차를 추출하지 못했습니다.")

        # 신규 포맷: 우측 요약 블록 + 생산 테이블이 있으면 production(투입/불량/손실) 산출.
        prod_rows = []
        prod = self._parse_new_production(ws, year, week, week_num, warnings)
        if prod is not None:
            prod_rows = [prod]

        production, detail = self.build_frames(prod_rows, detail_rows)
        return ParseResult(production=production, defect_detail=detail, warnings=warnings)

    def _parse_new_production(self, ws, year, week, week_num, warnings):
        """신규 Kopac 파일의 우측 요약 블록 + 생산 테이블에서 투입/불량/손실 계산.

        - 요약 블록(c12 라벨, c13 값): 불량유형별 합계(공정검사 포함), 마지막에 총합.
        - 생산 테이블(헤더 '양품수량'·'기간합계'): 투입 = Σ(기간합계 양품) + 불량합계.
        - 불량 = 불량합계 − 공정검사(손실),  손실 = 공정검사.
        반환: production dict (없으면 None).
        """
        # 1) 요약 블록 읽기 (c12=라벨, c13=값)
        summary, total = {}, None
        for r in range(2, 30):
            lab = ws.cell(row=r, column=12).value
            val = ws.cell(row=r, column=13).value
            if lab and str(lab).strip():
                summary[str(lab).strip()] = normalize.clean_number(val)
            elif val is not None and summary:   # 라벨 없이 값만 → 총합
                total = normalize.clean_number(val)
                break
        if not summary:
            return None   # 구포맷(요약블록 없음) → production 미산출

        defect_total = total if total is not None else sum(summary.values())
        loss_qty = summary.get("공정검사", 0.0)
        defect_qty = defect_total - loss_qty

        # 2) 생산 테이블에서 '기간합계' 양품 합 → 투입
        good_qty = self._sum_period_good(ws)
        if good_qty is None:
            warnings.append("Kopac: 생산 테이블(기간합계 양품)을 찾지 못해 투입수량 미산출.")
            input_qty = 0.0
        else:
            input_qty = good_qty + defect_total   # 투입 = 양품 + (불량+손실) 합계

        return {
            "year": year, "week": week, "week_num": week_num, "date": None,
            "process": "Kopac", "model": None,
            "input_qty": input_qty, "defect_qty": defect_qty, "loss_qty": loss_qty,
            "defect_rate": normalize.safe_rate(defect_qty, input_qty),
            "loss_rate": normalize.safe_rate(loss_qty, input_qty),
            "inspection_included": False, "setting_included": True,
        }

    @staticmethod
    def _sum_period_good(ws):
        """생산 테이블의 '기간합계' 블록 양품수량을 제품행 합산해 반환(없으면 None)."""
        # '기간합계' 라벨이 있는 헤더 행/열 탐색
        gc = gr = None
        for r in range(20, min(40, ws.max_row) + 1):
            for c in range(12, min(40, ws.max_column) + 1):
                if str(ws.cell(row=r, column=c).value).strip() == "기간합계":
                    gr, gc = r, c
                    break
            if gc:
                break
        if gc is None:
            return None
        # 기간합계 블록 안에서 '양품수량' 헤더 열 찾기(보통 다음 행에 헤더)
        good_col = None
        for hr in (gr, gr + 1):
            for c in range(gc, min(gc + 5, ws.max_column) + 1):
                if str(ws.cell(row=hr, column=c).value).strip() == "양품수량":
                    good_col = c
                    hdr_row = hr
                    break
            if good_col:
                break
        if good_col is None:
            return None
        # 표의 좌측 키 열(총계획수량/생산지시수량) — 제품 행에만 값이 있고 합계 행은 비어 있음
        key_col = None
        for c in range(10, gc):
            if str(ws.cell(row=hdr_row, column=c).value).strip() in ("총계획수량", "생산지시수량"):
                key_col = c
                break
        # 제품 행만 합산(키 열이 빈 합계행에서 중단)
        s = 0.0
        for r in range(hdr_row + 1, ws.max_row + 1):
            if key_col is not None:
                if ws.cell(row=r, column=key_col).value in (None, ""):
                    break
            elif not any(ws.cell(row=r, column=c).value not in (None, "") for c in range(12, gc)):
                break
            s += normalize.clean_number(ws.cell(row=r, column=good_col).value)
        return s

    # ── (2) 전수검사 ───────────────────────────────────────────────
    def _parse_inspection(self, wb, year, week, week_num) -> ParseResult:
        ws = wb.worksheets[0]
        warnings: list[str] = []
        prod_rows, detail_rows = [], []
        TYPE_COL, NAME_COL, SUM_COL = 3, 2, 6
        nrow = ws.max_row

        # 1) 셀 캐시 + 섹션 헤더('…전수검사')·블록 시작('검사수량') 위치 수집
        names, types, vals = {}, {}, {}
        sections, starts = [], []
        for r in range(1, nrow + 1):
            n = ws.cell(row=r, column=NAME_COL).value
            t = ws.cell(row=r, column=TYPE_COL).value
            v = ws.cell(row=r, column=SUM_COL).value
            names[r], types[r], vals[r] = n, t, v
            if n and "전수검사" in str(n):
                sections.append((r, str(n).strip()))
            if t and str(t).strip() == "검사수량":
                starts.append(r)

        def section_of(row):
            sec = ""
            for sr, title in sections:
                if sr <= row:
                    sec = title
            return sec

        # 2) 블록별 파싱
        for idx, rs in enumerate(starts):
            re_ = starts[idx + 1] - 1 if idx + 1 < len(starts) else nrow
            # 다음 섹션 헤더가 블록 중간에 오면 거기서 끊음
            for sr, _t in sections:
                if rs < sr <= re_:
                    re_ = sr - 1
                    break
            is_bulk = "bulk" in section_of(rs).lower() or "벌크" in section_of(rs)
            model = self._model_of(names[rs], is_bulk)

            block, type_rows = {}, []
            for r in range(rs, re_ + 1):
                label = str(types[r]).strip() if types[r] is not None else ""
                if not label:
                    continue
                val = normalize.clean_number(vals[r])
                if label in ("검사수량", "적합수량", "부적합수량", "공정검사수량", "불량율", "부적합"):
                    block[label] = val
                else:
                    type_rows.append((label, val))

            input_qty = block.get("검사수량", 0.0)
            defect_qty = block.get("부적합수량", block.get("부적합", 0.0))
            loss_qty = block.get("공정검사수량", 0.0)
            prod_rows.append({
                "year": year, "week": week, "week_num": week_num, "date": None,
                "process": "Kopac", "model": model,
                "input_qty": input_qty, "defect_qty": defect_qty, "loss_qty": loss_qty,
                "defect_rate": normalize.safe_rate(defect_qty, input_qty),
                "loss_rate": normalize.safe_rate(loss_qty, input_qty),
                "inspection_included": False, "setting_included": True,
            })
            for label, val in type_rows:
                if val == 0:
                    continue
                detail_rows.append({
                    "year": year, "week": week, "week_num": week_num, "date": None,
                    "process": "Kopac", "model": model,
                    "defect_type": normalize.standardize_defect_type("Kopac", label),
                    "defect_qty": val, "is_loss": False,
                })
            if loss_qty:
                detail_rows.append({
                    "year": year, "week": week, "week_num": week_num, "date": None,
                    "process": "Kopac", "model": model,
                    "defect_type": "공정검사", "defect_qty": loss_qty, "is_loss": True,
                })

        if not prod_rows:
            warnings.append("Kopac 전수검사: 제품 블록을 찾지 못했습니다.")
        prod, detail = self.build_frames(prod_rows, detail_rows)
        return ParseResult(production=prod, defect_detail=detail, warnings=warnings)

    @staticmethod
    def _model_of(product_name, is_bulk=False) -> str:
        s = str(product_name)
        if "1ml" in s:
            return "1ml"
        if is_bulk:
            return "3ml bulk"
        if "3ml" in s:
            return "3ml"
        return s.strip()
