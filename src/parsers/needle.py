# -*- coding: utf-8 -*-
"""주사침 공정 파서.

입력: '2026-W##-불량유형_주사침 MMDD.xlsx' 의 'Sheet1'.
구조(요약):
  - 상단 생산현황 표: 헤더(W14..W17 | 일별 | 합계), 행=계획수량/모델들/적합수량/불량수/공정소모/불량율
  - 하단 불량유형 표: 헤더(W14..W17 | 일별 | 불량합계),
      블록='주 사 침 불 량' / '안 과 용·범 용 불 량' / '의 료 기 기 불 량'
  - col14(합계/불량합계) = 당주(W18) 일별 합계  ← 우리가 쓰는 현재주차 값

요청사항 적용:
  - 불량률은 '주사침 / 범용 / 안과용'만 관리. '의료기기 set 모델'(config.NEEDLE_EXCLUDE_MODELS)은 제외.
  - 세팅소모 / QC검사 / 공정검사는 손실(loss)로 분리(요청 5번).
입력수량 = (관리대상 모델 적합 합) + (관리대상 불량 합)  ← 파일의 불량율 정의(적합+불량)와 일치.
"""
from __future__ import annotations
import re

import openpyxl

from .base import BaseParser, ParseResult
from .. import normalize, config

_LOSS = {re.sub(r"\s+", "", x) for x in config.LOSS_TYPES["주사침"]}      # {세팅소모,QC검사,공정검사}
_EXCLUDE = {re.sub(r"\s+", "", x) for x in config.NEEDLE_EXCLUDE_MODELS}
_TOTALS = {"총불량수량", "총로스수량"}


def _ns(s) -> str:
    return re.sub(r"\s+", "", str(s)) if s is not None else ""


def _category(model_name: str) -> str | None:
    """모델명 → 카테고리(주사침/범용/안과용). 제외 모델이면 None."""
    key = _ns(model_name)
    if key in _EXCLUDE:
        return None
    s = str(model_name)
    if "안과" in s:
        return "안과용"
    if "범용" in s:
        return "범용"
    return "주사침"   # 멸균주사침 / EZ-injec 등


class NeedleParser(BaseParser):
    process = "주사침"

    def parse(self, file, week_label: str | None = None) -> ParseResult:
        warnings: list[str] = []
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.worksheets[0]

        # 1) 현재 주차: 명시값 > 시트 텍스트(W##) > 파일명
        year = week = week_num = None
        if week_label:
            year, week, week_num = normalize.normalize_week(week_label)
        else:
            for row in ws.iter_rows(min_row=1, max_row=12, max_col=12, values_only=True):
                for v in row:
                    if v and re.search(r"W\s*\d{1,2}", str(v)):
                        try:
                            year, week, week_num = normalize.normalize_week(v)
                            break
                        except ValueError:
                            continue
                if week:
                    break
            if week is None:
                d = normalize.yymmdd_from_name(getattr(file, "name", str(file)))
                if d:
                    year, week, week_num = normalize.week_from_date(d)

        # 2) 열/행 앵커 탐색
        # 우측에 별도 '게이지별 불량' 표가 있으므로 메인 표(col 2~14)로 한정해 탐색.
        prod_total_col = self._find_label_col(ws, "합계", cmax=14)
        defect_total_col = self._find_label_col(ws, "불량합계", cmax=14)
        plan_row = self._find_label_row(ws, "계획수량", 3)
        fit_row = self._find_label_row(ws, "적합수량", 3)
        defect_hdr_row = self._find_label_row(ws, "불량합계", None, cmax=14)  # 불량유형 헤더 행

        if None in (prod_total_col, defect_total_col, fit_row):
            warnings.append("주사침: 핵심 앵커(합계열/적합수량행)를 찾지 못했습니다.")
            wb.close()
            return ParseResult(*self.build_frames([], []), warnings=warnings)

        # 3) 관리대상 모델 적합수량 합 (의료기기 제외)
        managed_fit = 0.0
        if plan_row and fit_row and plan_row < fit_row:
            for r in range(plan_row + 1, fit_row):
                name = ws.cell(row=r, column=3).value
                if not name:
                    continue
                cat = _category(name)
                if cat is None:        # 의료기기 set 모델 → 제외
                    continue
                managed_fit += normalize.clean_number(ws.cell(row=r, column=prod_total_col).value)

        # 4) 불량유형 표 파싱 (블록 추적)
        detail_rows = []
        managed_defect = 0.0
        loss_qty = 0.0
        excluded_defect = 0.0
        current_block = None
        start = (defect_hdr_row or 43) + 1
        for r in range(start, ws.max_row + 1):
            blk = ws.cell(row=r, column=2).value
            if blk:
                current_block = _ns(blk)
            typ = ws.cell(row=r, column=3).value
            if not typ:
                continue
            key = _ns(typ)
            if key in _TOTALS:
                continue
            qty = normalize.clean_number(ws.cell(row=r, column=defect_total_col).value)

            if key in _LOSS:                              # 세팅소모/QC검사/공정검사 → 손실
                loss_qty += qty
                if qty:
                    detail_rows.append(self._d(year, week, week_num, "주사침",
                                               normalize.standardize_defect_type("주사침", typ),
                                               qty, True))
                continue
            if (current_block and "의료기기" in current_block) or key in _EXCLUDE:
                excluded_defect += qty                    # 의료기기 set → 제외
                continue
            # 관리대상 불량(주사침/안과·범용)
            managed_defect += qty
            cat = "안과·범용" if (current_block and "안" in current_block) else "주사침"
            if qty:
                detail_rows.append(self._d(year, week, week_num, cat,
                                           normalize.standardize_defect_type("주사침", typ),
                                           qty, False))

        # 5) 생산 그레인 (관리대상 합산)
        input_qty = managed_fit + managed_defect
        prod_rows = [{
            "year": year, "week": week, "week_num": week_num, "date": None,
            "process": "주사침", "model": None,
            "input_qty": input_qty, "defect_qty": managed_defect, "loss_qty": loss_qty,
            "defect_rate": normalize.safe_rate(managed_defect, input_qty),
            "loss_rate": normalize.safe_rate(loss_qty, input_qty),
            "inspection_included": False, "setting_included": False,
        }]
        prod, detail = self.build_frames(prod_rows, detail_rows)
        wb.close()
        # 진단용: 제외된 의료기기 불량량을 경고에 남김(정보)
        if excluded_defect:
            warnings.append(f"주사침: 의료기기 set 불량 {excluded_defect:.0f}건 제외(별도 관리).")
        return ParseResult(production=prod, defect_detail=detail, warnings=warnings)

    # ── 보조 ───────────────────────────────────────────────────────
    def _d(self, year, week, week_num, model, dtype, qty, is_loss):
        return {"year": year, "week": week, "week_num": week_num, "date": None,
                "process": "주사침", "model": model,
                "defect_type": dtype, "defect_qty": qty, "is_loss": is_loss}

    @staticmethod
    def _find_label_col(ws, label, cmax=14, rstart=15, rend=50):
        for r in range(rstart, min(rend, ws.max_row) + 1):
            for c in range(2, min(cmax, ws.max_column) + 1):
                if _ns(ws.cell(row=r, column=c).value) == _ns(label):
                    return c
        return None

    @staticmethod
    def _find_label_row(ws, label, col, rmax=92, cmax=14):
        rng = range(1, min(rmax, ws.max_row) + 1)
        cols = [col] if col else range(2, min(cmax, ws.max_column or 2) + 1)
        for r in rng:
            for c in cols:
                if _ns(ws.cell(row=r, column=c).value) == _ns(label):
                    return r
        return None
