# -*- coding: utf-8 -*-
"""Kopac 파서 검증.

(A) 전수검사: 부적합 = 검사-적합, 유형별 합 = 부적합, 공정검사(손실) 분리,
    파일명(YYMMDD)→주차 추출 확인.
(B) 파레토: 1ml/3ml 분리, '공정검사' 포함 불량명을 손실로 분류,
    재계산한 점유율이 파일의 점유율(%)과 일치.
"""
from __future__ import annotations
import sys
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from src.parsers.kopac import KopacParser  # noqa: E402

KO = Path(r"D:/Project/불량률/생산 현황 및 불량/Kopac")
INSP = KO / "전수검사 내역_코팩_260427.xlsx"
PARETO = KO / "불량유형_코팩_260504.xlsx"
TOL = 0.01  # 점유율(%) 허용오차


def run():
    fails = []

    # ── (A) 전수검사 ──────────────────────────────────────────────
    print("=== (A) 전수검사 ===")
    res = KopacParser().parse(INSP)
    prod, detail = res.production, res.defect_detail
    for w in res.warnings:
        print("   ⚠", w)
    print(f"파싱 production {len(prod)}행:")
    print(prod[["week", "model", "input_qty", "defect_qty", "loss_qty", "defect_rate"]].to_string(index=False))

    # 파일명 260427 → ISO 주차 18 기대
    if not (prod["week"] == "26-W18").all():
        fails.append(f"주차 추출 오류: {prod['week'].unique()}")

    # 1ml 검산: 검사 529130, 적합 526500 → 부적합 2630, 공정검사 4062
    one = prod[prod["model"] == "1ml"].iloc[0]
    if abs(one["input_qty"] - 529130) > 0.5: fails.append("1ml 검사수량")
    if abs(one["defect_qty"] - 2630) > 0.5: fails.append(f"1ml 부적합 {one['defect_qty']}")
    if abs(one["loss_qty"] - 4062) > 1.5: fails.append(f"1ml 공정검사(손실) {one['loss_qty']}")
    # 유형별 합 = 부적합(손실 제외분)
    d1 = detail[(detail["model"] == "1ml") & (~detail["is_loss"])]["defect_qty"].sum()
    print(f"\n1ml 유형별 합(손실제외)={d1:.0f} vs 부적합={one['defect_qty']:.0f}")
    if abs(d1 - one["defect_qty"]) > 1.5:
        fails.append(f"1ml 유형합≠부적합 ({d1:.0f}≠{one['defect_qty']:.0f})")

    # 3ml(완제품)·3ml bulk 분리 확인
    models = set(prod["model"])
    print("model 종류:", models)
    if "3ml bulk" not in models:
        fails.append("3ml bulk 블록 미분리")

    # ── (B) 파레토 ────────────────────────────────────────────────
    print("\n=== (B) 파레토 ===")
    resp = KopacParser().parse(PARETO)
    pdet = resp.defect_detail
    for w in resp.warnings:
        print("   ⚠", w)
    print(f"파싱 defect_detail {len(pdet)}행, model={set(pdet['model'])}, 주차={set(pdet['week'])}")

    # 손실(공정검사) 분류 확인
    losses = pdet[pdet["is_loss"]]
    print("손실로 분류된 항목:")
    print(losses[["model", "defect_type", "defect_qty"]].to_string(index=False))
    if losses.empty:
        fails.append("파레토: 공정검사 손실 분류 0건")

    # 점유율 재계산 검증: 파일 r3 1ml S025(본딩공정검사)=270, 점유율 25.6653%
    wb = openpyxl.load_workbook(PARETO, data_only=True)
    ws = wb[[s for s in wb.sheetnames if s.strip() == "Defect Repair Summary By Type"][0]]
    file_top_qty = ws.cell(row=3, column=3).value      # 270
    file_top_share = ws.cell(row=3, column=4).value    # 25.6653 (%)
    one_ml = pdet[pdet["model"] == "1ml"]
    total_1ml = one_ml["defect_qty"].sum()
    recomputed_share = file_top_qty / total_1ml * 100
    print(f"\n1ml 최상위 점유율 재계산: {recomputed_share:.4f}% vs 파일 {file_top_share:.4f}%")
    if abs(recomputed_share - file_top_share) > TOL:
        fails.append(f"점유율 불일치 {recomputed_share:.4f} vs {file_top_share:.4f}")
    wb.close()

    # ── (C) 신규 파일(W25) 생산량 → 마스터 회귀 ──────────────────
    print("\n=== (C) 신규 파일 W25 생산량 vs 마스터 ===")
    NEW = KO / "불량유형_코팩_2606022.xlsx"
    if NEW.exists():
        rn = KopacParser().parse(NEW)
        for w in rn.warnings:
            print("   ⚠", w)
        if rn.production.empty:
            fails.append("신규파일 production 미산출")
        else:
            pr = rn.production.iloc[0]
            print(pr[["week", "input_qty", "defect_qty", "loss_qty", "defect_rate"]].to_string())
            # 마스터 W25 Kopac: 투입 235836 / 불량 852 / 불량률 0.003613
            for nm, got, exp, tol in [("투입", pr["input_qty"], 235836, 0.5),
                                       ("불량", pr["defect_qty"], 852, 0.5),
                                       ("손실(공정검사)", pr["loss_qty"], 1284, 0.5),
                                       ("불량률", pr["defect_rate"], 0.003612679997964687, 1e-4)]:
                ok = abs(got - exp) <= tol
                print(f"  {nm}: 파싱 {got} / 마스터 {exp}  {'OK' if ok else '✗'}")
                if not ok:
                    fails.append(f"신규W25 {nm}")
            if pr["week"] != "26-W25":
                fails.append(f"신규W25 주차 {pr['week']}")

    print("\n" + ("✅ Kopac 파서 검증 통과" if not fails else f"❌ 불일치 {fails}"))
    return not fails


if __name__ == "__main__":
    sys.exit(0 if run() else 1)
