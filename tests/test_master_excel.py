# -*- coding: utf-8 -*-
"""마스터 취합본 재생성 검증.

build_master 로 만든 xlsx 를 다시 읽어,
'주차별 생산현황' 시트의 W25 공정별 값과 우측 평균/가중 불량률이
원본 마스터와 4자리 일치하는지 확인.
"""
from __future__ import annotations
import sys
import io
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from src import pipeline, master_excel  # noqa: E402

BASE = r"D:/Project/불량률/생산 현황 및 불량"
MASTER_W25 = {"avg": 0.003250530248025093, "weighted": 0.002390705403868861}
EXPECT_PROC = {  # W25 공정별 (투입, 불량)
    "Kopac": (235836, 852), "주사침": (429057, 2501),
    "사출 통합": (2837421, 5251), "연마": (339230, 580),
}
TOL = 1e-4


def run():
    load = pipeline.load_from_folder(BASE)
    data = master_excel.build_master(load.production, load.defect_detail)
    assert data and len(data) > 1000, "xlsx 바이트가 비정상"

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    print("생성된 시트:", wb.sheetnames)
    fails = []
    need = {"주차별 생산현황", "모델별 불량률 현황", "Raw_production", "Raw_defect_detail"}
    missing = need - set(wb.sheetnames)
    if missing:
        fails.append(f"누락 시트 {missing}")

    ws = wb["주차별 생산현황"]
    # 좌측 블록: W25 공정별 투입/불량 수집
    got = {}
    for r in range(2, ws.max_row + 1):
        wk, proc = ws.cell(r, 3).value, ws.cell(r, 4).value
        if wk == "26-W25" and proc:
            got[proc] = (ws.cell(r, 5).value, ws.cell(r, 6).value)
    print("\n[생성본 W25 좌측 블록]")
    for proc, (inp, dfq) in EXPECT_PROC.items():
        g = got.get(proc)
        ok = g and abs(g[0] - inp) < 0.5 and abs(g[1] - dfq) < 0.5
        print(f"  {proc}: {g} / 기대 ({inp}, {dfq})  {'OK' if ok else '✗'}")
        if not ok:
            fails.append(f"좌측 {proc}")

    # 우측 요약: W25 평균/가중
    avg = wgt = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(r, 9).value == "26-W25":
            avg, wgt = ws.cell(r, 10).value, ws.cell(r, 11).value
            break
    print(f"\n[생성본 W25 우측 요약] 평균 {avg} / 가중 {wgt}")
    if avg is None or abs(avg - MASTER_W25["avg"]) > TOL:
        fails.append("우측 평균")
    if wgt is None or abs(wgt - MASTER_W25["weighted"]) > TOL:
        fails.append("우측 가중")

    print("\n" + ("✅ 마스터 재생성 검증 통과 — W25 좌/우 블록 마스터 일치"
                  if not fails else f"❌ 불일치 {fails}"))
    return not fails


if __name__ == "__main__":
    sys.exit(0 if run() else 1)
